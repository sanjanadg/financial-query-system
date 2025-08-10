import pandas as pd
import openpyxl
import numpy as np
from typing import Dict, List, Any, Tuple, Optional
import re
from datetime import datetime
import warnings
from .hierarchical_graph import HierarchicalGraphBuilder
from .knowledge_graph import KnowledgeGraphBuilder
warnings.filterwarnings('ignore')

"""
ENHANCED DATA REPRESENTATION STRUCTURE:
"""

class AtomicUnit:
    """Represents an atomic unit of data with localization and context."""
    
    def __init__(self, unit_type: str, content: Any, location: Dict, context: Dict = None):
        self.unit_type = unit_type  # 'cell', 'row', 'column', 'table', 'sheet'
        self.content = content      # The actual data content
        self.location = location    # Where this unit is located
        self.context = context or {}  # Contextual information
        self.hierarchy = {}         # Hierarchical relationships
        self.embedding = None       # Will be computed later
        
    def add_context(self, key: str, value: Any):
        """Add contextual information to this atomic unit."""
        self.context[key] = value
        
    def add_hierarchy(self, parent: str, child: str):
        """Add hierarchical relationship."""
        if parent not in self.hierarchy:
            self.hierarchy[parent] = []
        self.hierarchy[parent].append(child)
        
    def get_full_description(self) -> str:
        """Generate a comprehensive description for embedding."""
        desc_parts = [
            f"Type: {self.unit_type}",
            f"Content: {str(self.content)}",
            f"Location: {self.location}"
        ]
        
        if self.context:
            context_str = ", ".join([f"{k}: {v}" for k, v in self.context.items()])
            desc_parts.append(f"Context: {context_str}")
            
        if self.hierarchy:
            hierarchy_str = ", ".join([f"{k}->{v}" for k, v in self.hierarchy.items()])
            desc_parts.append(f"Hierarchy: {hierarchy_str}")
            
        return " | ".join(desc_parts)

class ExcelDataRepresentation:
    """A comprehensive representation of Excel financial data following the 5-step approach."""
    
    def __init__(self):
        # Step 1: Direct integration with data sources
        self.source_info = {
            'file_path': None,
            'file_type': 'excel',
            'sheets': [],
            'last_modified': None
        }
        
        # Step 2: Atomic units localization
        self.atomic_units = {
            'cells': [],
            'rows': [],
            'columns': [],
            'tables': [],
            'sheets': []
        }
        
        # Step 3: Contextual information and hierarchy
        self.contextual_info = {
            'company_hierarchy': {},
            'time_hierarchy': {},
            'metric_hierarchy': {},
            'formula_dependencies': {},
            'data_relationships': {}
        }
        
        # Step 4: Embeddings (will be computed)
        self.embeddings = {}
        
        # Step 5: Graph representation
        self.knowledge_graph = None
        
        # Legacy support
        self.sheets_data = {}
        self.named_ranges = {}
        self.formulas = {}
        self.company_data = {}
        self.monthly_data = {}
        self.debt_schedules = {}
        self.metadata = {}
        
    def add_atomic_unit(self, unit: AtomicUnit):
        """Add an atomic unit to the appropriate category."""
        if unit.unit_type in self.atomic_units:
            self.atomic_units[unit.unit_type].append(unit)
        else:
            self.atomic_units[unit.unit_type] = [unit]
            
    def add_contextual_info(self, category: str, key: str, value: Any):
        """Add contextual information."""
        if category not in self.contextual_info:
            self.contextual_info[category] = {}
        self.contextual_info[category][key] = value
        
    def get_all_atomic_units(self) -> List[AtomicUnit]:
        """Get all atomic units across all categories."""
        all_units = []
        for units in self.atomic_units.values():
            all_units.extend(units)
        return all_units
        
    def add_sheet_data(self, sheet_name: str, data: pd.DataFrame, formulas: Dict = None):
        """Add processed sheet data to the representation."""
        self.sheets_data[sheet_name] = data
        if formulas:
            self.formulas[sheet_name] = formulas

class ExcelProcessor:
    """Processes Excel files following the 5-step data representation approach."""
    
    def __init__(self):
        self.companies = ['MXD', 'HEC', 'Branch']
        self.months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 
                      'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        self.years = [2020, 2021, 2022, 2023, 2024]
        
        # Step 1: Data source integration capabilities
        self.supported_formats = ['xlsx', 'xlsm', 'xls']
        self.data_sources = {
            'excel': self._process_excel_source,
            # Future: 'api': self._process_api_source,
            # Future: 'database': self._process_database_source
        }
        
    def extract_cell_formulas(self, worksheet) -> Dict[str, str]:
        """Extract formulas from worksheet cells."""
        formulas = {}
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formulas[f"{cell.column_letter}{cell.row}"] = cell.value
        return formulas
    
    def clean_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """Clean and standardize dataframe."""
        # Remove completely empty rows and columns
        df = df.dropna(how='all').dropna(axis=1, how='all')
        
        # Fill NaN with 0 for numeric columns
        numeric_cols = df.select_dtypes(include=[np.number]).columns
        df[numeric_cols] = df[numeric_cols].fillna(0)
        
        return df
    
    def extract_monthly_data(self, df: pd.DataFrame, sheet_name: str) -> Dict:
        """Extract monthly financial data from dataframe."""
        monthly_data = {}
        
        # Check if first row contains month/year headers
        if len(df.columns) > 1 and len(df) > 0:
            first_row = df.iloc[0]
            
            # Look for month/year patterns in the first row
            for col_idx, col_value in enumerate(first_row):
                if pd.notna(col_value) and isinstance(col_value, str):
                    col_str = str(col_value).strip()
                    
                    # Parse month and year from header
                    for month in self.months:
                        if month.lower() in col_str.lower():
                            # Extract year from the header
                            for year in self.years:
                                if str(year) in col_str:
                                    key = f"{month}_{year}"
                                    if key not in monthly_data:
                                        monthly_data[key] = {}
                                    
                                    # Extract data from this column
                                    for row_idx in range(1, len(df)):
                                        if row_idx < len(df):
                                            row_name = str(df.iloc[row_idx, 0]) if len(df.columns) > 0 else f"row_{row_idx}"
                                            cell_value = df.iloc[row_idx, col_idx] if col_idx < len(df.columns) else None
                                            
                                            if pd.notna(cell_value) and isinstance(cell_value, (int, float)):
                                                monthly_data[key][row_name] = cell_value
                                    break
        
        return monthly_data
    
    def extract_company_data(self, df: pd.DataFrame, sheet_name: str) -> Dict:
        """Extract company-specific data."""
        company_data = {}
        
        for company in self.companies:
            company_data[company] = {}
            
            # Look for rows containing company name
            for idx, row in df.iterrows():
                if idx == 0:  # Skip header row
                    continue
                    
                row_str = ' '.join([str(cell) for cell in row if pd.notna(cell)]).lower()
                
                # Check if company is mentioned
                if company.lower() in row_str:
                    # Extract associated data
                    for col_idx, col in enumerate(df.columns):
                        if pd.notna(row[col]) and isinstance(row[col], (int, float)):
                            col_name = str(col) if col else f"col_{col_idx}"
                            company_data[company][col_name] = row[col]
        
        return company_data
    
    def extract_debt_schedules(self, workbook) -> Dict:
        """Extract debt schedule information."""
        debt_data = {}
        
        debt_sheets = [sheet for sheet in workbook.sheetnames if 'debt' in sheet.lower()]
        
        for sheet_name in debt_sheets:
            worksheet = workbook[sheet_name]
            df = pd.DataFrame(worksheet.values)
            df = self.clean_dataframe(df)
            
            debt_data[sheet_name] = {
                'data': df,
                'formulas': self.extract_cell_formulas(worksheet)
            }
        
        return debt_data
    
    def process_sheet(self, worksheet, sheet_name: str) -> Dict:
        """Process individual worksheet and extract structured data."""
        # Convert to dataframe
        df = pd.DataFrame(worksheet.values)
        df = self.clean_dataframe(df)
        
        # Extract formulas
        formulas = self.extract_cell_formulas(worksheet)
        
        # Extract monthly data
        monthly_data = self.extract_monthly_data(df, sheet_name)
        
        # Extract company data
        company_data = self.extract_company_data(df, sheet_name)
        
        return {
            'data': df,
            'formulas': formulas,
            'monthly_data': monthly_data,
            'company_data': company_data,
            'sheet_name': sheet_name
        }
    
    def process_file(self, file_path: str) -> ExcelDataRepresentation:
        """Process Excel file following the 5-step data representation approach."""
        print(f"Processing file: {file_path}")
        
        # Step 1: Integrate directly into data sources
        representation = self._integrate_data_source(file_path)
        if not representation:
            return None
            
        # Step 2: Localize each atomic unit
        print("Step 2: Localizing atomic units...")
        self._localize_atomic_units(representation)
        
        # Step 3: Compute contextual information and hierarchy
        print("Step 3: Computing contextual information and hierarchy...")
        self._compute_context_and_hierarchy(representation)
        
        # Step 4: Construct embeddings using content and context
        print("Step 4: Constructing embeddings...")
        self._construct_embeddings(representation)
        
        # Step 5: Insert atomic units into graph representation
        print("Step 5: Building hierarchical knowledge graph...")
        representation.knowledge_graph = self.build_hierarchical_knowledge_graph(representation)
        
        print(f"Processing complete. Extracted data from {len(representation.source_info['sheets'])} sheets.")
        return representation
        
    def _integrate_data_source(self, file_path: str) -> ExcelDataRepresentation:
        """Step 1: Integrate directly into data sources."""
        print("Step 1: Integrating data source...")
        
        # Determine data source type
        file_extension = file_path.split('.')[-1].lower()
        if file_extension in self.supported_formats:
            source_type = 'excel'
        else:
            print(f"Unsupported file format: {file_extension}")
            return None
            
        # Initialize representation with source info
        representation = ExcelDataRepresentation()
        representation.source_info.update({
            'file_path': file_path,
            'file_type': source_type,
            'last_modified': datetime.now().isoformat()
        })
        
        # Process the data source
        if source_type in self.data_sources:
            return self.data_sources[source_type](file_path, representation)
        else:
            print(f"Unsupported data source type: {source_type}")
            return None
            
    def _process_excel_source(self, file_path: str, representation: ExcelDataRepresentation) -> ExcelDataRepresentation:
        """Process Excel data source."""
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=False)
            representation.source_info['sheets'] = workbook.sheetnames
            
            # Process each sheet
            for sheet_name in workbook.sheetnames:
                print(f"Processing sheet: {sheet_name}")
                worksheet = workbook[sheet_name]
                
                # Convert to DataFrame
                data = []
                for row in worksheet.iter_rows(values_only=True):
                    data.append(row)
                
                df = pd.DataFrame(data)
                df = self.clean_dataframe(df)
                
                # Extract formulas
                formulas = self.extract_cell_formulas(worksheet)
                
                # Process sheet data
                sheet_data = self.process_sheet(worksheet, sheet_name)
                
                # Add to representation (legacy support)
                representation.add_sheet_data(sheet_name, df, formulas)
                representation.monthly_data.update(sheet_data.get('monthly_data', {}))
                representation.company_data.update(sheet_data.get('company_data', {}))
            
            # Extract debt schedules
            representation.debt_schedules = self.extract_debt_schedules(workbook)
            
            return representation
            
        except Exception as e:
            print(f"Error processing Excel source: {e}")
            return None
            
    def _localize_atomic_units(self, representation: ExcelDataRepresentation):
        """Step 2: Localize each atomic unit (cells, rows, tables, etc.)."""
        print("Localizing atomic units...")
        
        for sheet_name, sheet_data in representation.sheets_data.items():
            df = sheet_data
            
            # Create sheet atomic unit
            sheet_unit = AtomicUnit(
                unit_type='sheet',
                content=sheet_name,
                location={'sheet_name': sheet_name},
                context={'total_rows': len(df), 'total_columns': len(df.columns)}
            )
            representation.add_atomic_unit(sheet_unit)
            
            # Create row atomic units
            for row_idx, row in df.iterrows():
                row_unit = AtomicUnit(
                    unit_type='row',
                    content=row.to_dict(),
                    location={'sheet_name': sheet_name, 'row_index': row_idx},
                    context={'row_type': 'data' if row_idx > 0 else 'header'}
                )
                representation.add_atomic_unit(row_unit)
                
                # Create cell atomic units
                for col_idx, (col_name, cell_value) in enumerate(row.items()):
                    if pd.notna(cell_value):
                        cell_unit = AtomicUnit(
                            unit_type='cell',
                            content=cell_value,
                            location={
                                'sheet_name': sheet_name,
                                'row_index': row_idx,
                                'column_index': col_idx,
                                'column_name': col_name
                            },
                            context={'data_type': type(cell_value).__name__}
                        )
                        representation.add_atomic_unit(cell_unit)
            
            # Create column atomic units
            for col_idx, col_name in enumerate(df.columns):
                col_data = df[col_name].dropna().tolist()
                if col_data:
                    col_unit = AtomicUnit(
                        unit_type='column',
                        content=col_data,
                        location={'sheet_name': sheet_name, 'column_name': col_name},
                        context={'data_count': len(col_data)}
                    )
                    representation.add_atomic_unit(col_unit)
                    
    def _compute_context_and_hierarchy(self, representation: ExcelDataRepresentation):
        """Step 3: Compute contextual information and hierarchy around and within each unit."""
        print("Computing contextual information and hierarchy...")
        
        # Analyze company hierarchy
        company_hierarchy = {}
        for company in self.companies:
            company_hierarchy[company] = {
                'subsidiaries': [],
                'departments': [],
                'metrics': []
            }
        representation.add_contextual_info('company_hierarchy', 'structure', company_hierarchy)
        
        # Analyze time hierarchy
        time_hierarchy = {
            'years': self.years,
            'quarters': ['Q1', 'Q2', 'Q3', 'Q4'],
            'months': self.months,
            'relationships': {
                'year_contains_quarters': {},
                'quarter_contains_months': {}
            }
        }
        representation.add_contextual_info('time_hierarchy', 'structure', time_hierarchy)
        
        # Analyze metric hierarchy
        metric_hierarchy = {
            'financial_metrics': ['revenue', 'profit', 'ebitda', 'fcf'],
            'operational_metrics': ['costs', 'expenses', 'labor'],
            'debt_metrics': ['debt', 'loan', 'liability']
        }
        representation.add_contextual_info('metric_hierarchy', 'structure', metric_hierarchy)
        
        # Analyze formula dependencies
        formula_deps = {}
        for sheet_name, formulas in representation.formulas.items():
            for cell_ref, formula in formulas.items():
                # Extract cell references from formula
                cell_refs = re.findall(r'[A-Z]+\d+', formula)
                formula_deps[cell_ref] = {
                    'sheet': sheet_name,
                    'formula': formula,
                    'dependencies': cell_refs
                }
        representation.add_contextual_info('formula_dependencies', 'structure', formula_deps)
        
        # Add hierarchy information to atomic units
        for unit in representation.get_all_atomic_units():
            if unit.unit_type == 'cell':
                # Add company context if cell contains company data
                for company in self.companies:
                    if company.lower() in str(unit.content).lower():
                        unit.add_context('company', company)
                        
                # Add time context if cell contains time data
                for month in self.months:
                    if month.lower() in str(unit.content).lower():
                        unit.add_context('time_period', month)
                        
                # Add metric context if cell contains metric data
                for metric_category, metrics in metric_hierarchy.items():
                    for metric in metrics:
                        if metric.lower() in str(unit.content).lower():
                            unit.add_context('metric_category', metric_category)
                            unit.add_context('metric', metric)
                            
    def _construct_embeddings(self, representation: ExcelDataRepresentation):
        """Step 4: Construct embeddings using both content and context."""
        print("Constructing embeddings...")
        
        # Initialize knowledge graph builder for embeddings
        kg_builder = KnowledgeGraphBuilder()
        
        # Create embeddings for each atomic unit
        for unit in representation.get_all_atomic_units():
            # Generate description for embedding
            description = unit.get_full_description()
            
            # Add node to knowledge graph (this will compute the embedding)
            node_id = f"{unit.unit_type}_{unit.location.get('sheet_name', 'unknown')}_{unit.location.get('row_index', 0)}_{unit.location.get('column_index', 0)}"
            
            kg_builder.add_node(
                node_id=node_id,
                node_type=unit.unit_type,
                node_data={
                    'content': unit.content,
                    'location': unit.location,
                    'context': unit.context,
                    'hierarchy': unit.hierarchy
                },
                description=description
            )
            
            # Store embedding reference
            representation.embeddings[node_id] = kg_builder.node_embeddings.get(node_id)
            
        print(f"Constructed embeddings for {len(representation.embeddings)} atomic units")
    
    def build_hierarchical_knowledge_graph(self, representation: ExcelDataRepresentation) -> HierarchicalGraphBuilder:
        """Build a hierarchical knowledge graph from the Excel data representation."""
        print("Building hierarchical knowledge graph...")
        
        # Use the new hierarchical graph builder
        hierarchical_builder = HierarchicalGraphBuilder()
        hierarchical_graph = hierarchical_builder.build_hierarchical_graph(representation)
        
        # Print summary
        summary = hierarchical_graph.get_graph_summary()
        print(f"Hierarchical knowledge graph built successfully:")
        print(f"  - Total nodes: {summary['total_nodes']}")
        print(f"  - Total edges: {summary['total_edges']}")
        print(f"  - Node types: {summary['node_types']}")
        print(f"  - Edge types: {summary['edge_types']}")
        print(f"  - Has embeddings: {summary['has_embeddings']}")
        print(f"  - Has search index: {summary['has_index']}")
        
        return hierarchical_graph
    
    # def build_knowledge_graph(self, representation: ExcelDataRepresentation) -> KnowledgeGraphBuilder:
    #     """Build a knowledge graph from the Excel data representation."""
    #     print("Initializing knowledge graph builder...")
    #     kg_builder = KnowledgeGraphBuilder()
        
    #     # Build company nodes
    #     print("Building company nodes...")
    #     company_nodes = kg_builder.build_company_nodes(self.companies)
        
    #     # Build metric nodes from the metrics mapping
    #     print("Building metric nodes...")
    #     metrics = [] 
    #     # Use the metrics from the query engine
    #     metrics_mapping = {
    #         'gross profit': ['gross profit', 'gross margin', 'gross income'],
    #         'shipping income': ['shipping income', 'shipping revenue', 'freight income'],
    #         'direct labor': ['direct labor', 'labor cost', 'direct labor cost'],
    #         'indirect costs': ['indirect costs', 'indirect expenses', 'overhead'],
    #         'operating expenses': ['operating expenses', 'opex', 'operational expenses'],
    #         'insurance': ['insurance', 'insurance expense', 'insurance cost'],
    #         'advertising': ['advertising', 'advertising expense', 'marketing'],
    #         'ebitda': ['ebitda', 'earnings before interest', 'operating profit'],
    #         'revenue': ['revenue', 'sales', 'income'],
    #         'fcf': ['fcf', 'free cash flow', 'cash flow'],
    #         'debt': ['debt', 'loan', 'liability']
    #     }
    #     for metric_group in metrics_mapping.values():
    #         metrics.extend(metric_group)
    #     metric_nodes = kg_builder.build_metric_nodes(metrics)
        
    #     # Build time period nodes
    #     print("Building time period nodes...")
    #     time_nodes = kg_builder.build_time_period_nodes(self.years, self.months)
        
    #     # Build sheet nodes
    #     print("Building sheet nodes...")
    #     sheet_nodes = kg_builder.build_sheet_nodes(representation.sheets_data)
        
    #     # Build data point nodes
    #     print("Building data point nodes...")
    #     data_nodes = kg_builder.build_data_point_nodes(representation.monthly_data)
        
    #     # Build formula nodes
    #     print("Building formula nodes...")
    #     all_formulas = {}
    #     for sheet_name, sheet_data in representation.sheets_data.items():
    #         if 'formulas' in sheet_data:
    #             all_formulas.update(sheet_data['formulas'])
    #     formula_nodes = kg_builder.build_formula_nodes(all_formulas)
        
    #     # Build relationships
    #     print("Building relationships...")
    #     kg_builder.build_relationships(
    #         company_nodes, metric_nodes, time_nodes, 
    #         sheet_nodes, data_nodes, formula_nodes
    #     )
        
    #     # Build search index
    #     print("Building search index...")
    #     kg_builder.build_index()
        
    #     # Print summary
    #     summary = kg_builder.get_graph_summary()
    #     print(f"Knowledge graph built successfully:")
    #     print(f"  - Total nodes: {summary['total_nodes']}")
    #     print(f"  - Total edges: {summary['total_edges']}")
    #     print(f"  - Has embeddings: {summary['has_embeddings']}")
    #     print(f"  - Has search index: {summary['has_index']}")
        
    #     return kg_builder

def process_file(file_path: str) -> Any:
    """
    Processes the Excel file and returns a structured representation.
    
    Args:
        file_path: Path to the Excel file
        
    Returns:
        ExcelDataRepresentation object containing structured data
    """
    processor = ExcelProcessor()
    return processor.process_file(file_path)
