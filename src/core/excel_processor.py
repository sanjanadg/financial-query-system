#!/usr/bin/env python3
"""
Enhanced Excel Processor for Financial Data Analysis
Handles structured financial data extraction and processing.
"""

import openpyxl
import json
import re
from sentence_transformers import SentenceTransformer
import numpy as np
from collections import defaultdict
import math

class ExcelProcessor:
    def __init__(self):
        """Initialize the processor with sentence transformer model."""
        self.model = SentenceTransformer('all-MiniLM-L6-v2')
        
    def process_file(self, file_path):
        """Process Excel file and create a comprehensive representation."""
        print(f"Processing file: {file_path}")
        
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            print(f"Error loading workbook: {e}")
            return None
        
        representation = {
            'file_path': file_path,
            'sheets': {},
            'financial_data': {},
            'metadata': {
                'total_sheets': len(workbook.sheetnames),
                'sheet_names': workbook.sheetnames
            }
        }
        
        for worksheet in workbook.worksheets:
            print(f"Processing sheet: {worksheet.title}")
            sheet_data = self._process_sheet(worksheet)
            if sheet_data:
                representation['sheets'][worksheet.title] = sheet_data
                
                # Extract financial data if this is a financial statement
                if 'financial_data' in sheet_data:
                    representation['financial_data'][worksheet.title] = sheet_data['financial_data']
        
        print(f"Successfully processed {len(workbook.worksheets)} sheets")
        return representation
    
    def _process_sheet(self, worksheet):
        """Process a single worksheet with enhanced financial data extraction."""
        try:
            # Get all data from the worksheet
            data = []
            for row in worksheet.iter_rows(values_only=True):
                data.append(row)
            
            if not data:
                return None
            
            # Identify financial structure
            financial_structure = self._identify_financial_structure(data)
            
            sheet_info = {
                'title': worksheet.title,
                'data': data,
                'rows': len(data),
                'columns': len(data[0]) if data else 0
            }
            
            # Extract financial data if structure is identified
            if financial_structure:
                financial_data = self._extract_financial_data(data, financial_structure)
                if financial_data:
                    sheet_info['financial_data'] = financial_data
                    sheet_info['financial_structure'] = financial_structure
            
            # Create embeddings for semantic search
            sheet_info['embeddings'] = self._create_embeddings(data)
            
            return sheet_info
            
        except Exception as e:
            print(f"Error processing sheet {worksheet.title}: {e}")
            return None
    
    def _identify_financial_structure(self, data):
        """Identify if the data represents a financial statement structure."""
        if not data:
            return None
            
        # Handle both list of lists and list of dictionaries
        if data and isinstance(data[0], dict):
            # Convert list of dictionaries to list of lists for processing
            converted_data = []
            for row in data:
                if isinstance(row, dict):
                    # Get all values in order
                    values = list(row.values())
                    converted_data.append(values)
                else:
                    converted_data.append(row)
            data = converted_data
        
        # For Branch P&L specifically, we know it's a financial statement
        # Check if this looks like Branch P&L by examining the data
        if len(data) > 0:
            # Look for any row that contains "final documentation" or similar Branch P&L indicators
            branch_indicators = ['final documentation', 'branch', 'brnch']
            has_branch_indicators = False
            for row in data[:20]:  # Check first 20 rows
                if row and len(row) > 0:
                    first_col_val = str(row[0]).lower() if row[0] else ''
                    if any(indicator in first_col_val for indicator in branch_indicators):
                        has_branch_indicators = True
                        break
            
            if has_branch_indicators:
                # Check if we have time periods in subsequent columns
                has_time_periods = False
                if len(data) > 0 and len(data[0]) > 1:
                    for col_idx in range(1, min(6, len(data[0]))):
                        if any(self._is_time_period(str(data[0][col_idx])) for row in data[:3]):
                            has_time_periods = True
                            break
                
                if has_time_periods:
                    return {
                        'type': 'financial_statement',
                        'special_case': 'branch_pl',
                        'has_time_periods': True
                    }
        
        # Standard financial statement detection
        if len(data) < 3:
            return None
        
        # Check first column for financial terms
        first_col_values = [str(row[0]).lower() if row and len(row) > 0 else '' 
                           for row in data[:10]]  # Check first 10 rows
        
        # Check if we have time periods in subsequent columns
        has_time_periods = False
        if len(data) > 0 and len(data[0]) > 1:
            for col_idx in range(1, min(6, len(data[0]))):
                if any(self._is_time_period(str(data[0][col_idx])) for row in data[:3]):
                    has_time_periods = True
                    break
        
        # Check for financial terms in first column
        financial_terms = ['income', 'revenue', 'gross profit', 'operating', 'ebitda', 
                          'cost', 'expense', 'profit', 'loss', 'cash', 'debt', 'equity',
                          'project', 'payment', 'milestone', 'documentation']
        
        has_financial_terms = any(any(term in val for term in financial_terms) 
                                 for val in first_col_values)
        
        if has_financial_terms and has_time_periods:
            return {
                'type': 'financial_statement',
                'has_time_periods': True
            }
        
        return None
    
    def _is_time_period(self, value):
        """Check if a value represents a time period."""
        if not value:
            return False
        
        value_str = str(value).lower().strip()
        
        # Month patterns
        month_patterns = [
            r'jan\s*\d{4}', r'feb\s*\d{4}', r'mar\s*\d{4}', r'apr\s*\d{4}',
            r'may\s*\d{4}', r'jun\s*\d{4}', r'jul\s*\d{4}', r'aug\s*\d{4}',
            r'sep\s*\d{4}', r'oct\s*\d{4}', r'nov\s*\d{4}', r'dec\s*\d{4}'
        ]
        
        # Year patterns
        year_patterns = [r'\d{4}', r'fy\d{2}', r'ttm\d{2}', r'ytd\d{2}']
        
        # Check month patterns
        for pattern in month_patterns:
            if re.match(pattern, value_str):
                return True
        
        # Check year patterns
        for pattern in year_patterns:
            if re.match(pattern, value_str):
                return True
        
        return False
    
    def _extract_financial_data(self, data, structure):
        """Extract financial data based on identified structure."""
        financial_data = {}
        
        # Handle both list of lists and list of dictionaries
        if data and isinstance(data[0], dict):
            # Convert list of dictionaries to list of lists for processing
            converted_data = []
            for row in data:
                if isinstance(row, dict):
                    # Get all values in order
                    values = list(row.values())
                    converted_data.append(values)
                else:
                    converted_data.append(row)
            data = converted_data
        
        # Handle special case for Branch P&L
        if structure.get('special_case') == 'branch_pl':
            return self._extract_branch_pl_data(data)
        
        # Standard financial data extraction
        if len(data) < 2:
            return financial_data
        
        # Extract headers (time periods)
        headers = data[0] if data else []
        time_periods = [str(h) for h in headers[1:] if h and self._is_time_period(str(h))]
        
        if not time_periods:
            return financial_data
        
        # Extract financial metrics from first column
        for row_idx, row in enumerate(data[1:], 1):
            if not row or len(row) == 0:
                continue
            
            metric_name = str(row[0]).strip() if row[0] else f"Row_{row_idx}"
            if not metric_name or metric_name.lower() in ['', 'nan', 'none']:
                continue
            
            # Extract values for each time period
            values = {}
            for col_idx, time_period in enumerate(time_periods, 1):
                if col_idx < len(row):
                    value = row[col_idx]
                    if value is not None and str(value).strip():
                        try:
                            # Clean and convert to numeric
                            if isinstance(value, str):
                                clean_value = value.replace('$', '').replace(',', '').replace('%', '')
                                if clean_value.replace('.', '').replace('-', '').isdigit():
                                    values[time_period] = float(clean_value)
                            elif isinstance(value, (int, float)):
                                values[time_period] = float(value)
                        except:
                            continue
            
            if values:  # Only add if we have actual values
                financial_data[metric_name] = {
                    'metric': metric_name,
                    'values': values,
                    'row_index': row_idx
                }
        
        return financial_data
    
    def _extract_branch_pl_data(self, data):
        """Extract financial data from Branch P&L with its special structure."""
        financial_data = {}
        
        # Branch P&L has a different structure - look for rows with numeric values
        for row_idx, row in enumerate(data):
            if not row or len(row) == 0:
                continue
            
            # Skip header rows (first 2 rows are typically headers)
            if row_idx < 2:
                continue
            
            # Look for rows that have numeric values in time period columns
            numeric_values = []
            time_periods = []
            
            for col_idx, value in enumerate(row):
                if col_idx == 0:  # Skip first column (labels)
                    continue
                
                if value is not None and str(value).strip():
                    try:
                        # Try to convert to numeric
                        if isinstance(value, str):
                            clean_value = value.replace('$', '').replace(',', '').replace('%', '')
                            if clean_value.replace('.', '').replace('-', '').isdigit():
                                numeric_value = float(clean_value)
                                numeric_values.append(numeric_value)
                                # Get the time period from the header
                                if len(data) > 0 and col_idx < len(data[0]):
                                    time_period = str(data[0][col_idx]).strip()
                                    time_periods.append(time_period)
                        elif isinstance(value, (int, float)):
                            numeric_values.append(float(value))
                            if len(data) > 0 and col_idx < len(data[0]):
                                time_period = str(data[0][col_idx]).strip()
                                time_periods.append(time_period)
                    except:
                        continue
            
            # If we found numeric values, create a metric
            if numeric_values and time_periods:
                # Create a descriptive metric name based on the row content
                metric_name = f"Branch Financial Metric Row {row_idx + 1}"
                
                # Try to get a meaningful name from the first column
                if row[0]:
                    first_col_val = str(row[0]).strip()
                    if first_col_val:
                        # Clean up the metric name
                        if first_col_val.lower() in ['final documentation', 'branch', 'brnch']:
                            # Skip these generic labels
                            continue
                        
                        # Use the first column value as the metric name
                        metric_name = f"Branch {first_col_val}"
                        
                        # Add context if it's a specific type of data
                        if any(term in first_col_val.lower() for term in ['forecast', 'budget', 'plan']):
                            metric_name = f"Branch {first_col_val} Forecast"
                        elif any(term in first_col_val.lower() for term in ['revenue', 'income', 'sales']):
                            metric_name = f"Branch {first_col_val}"
                        elif any(term in first_col_val.lower() for term in ['cost', 'expense', 'expenditure']):
                            metric_name = f"Branch {first_col_val}"
                        elif any(term in first_col_val.lower() for term in ['profit', 'margin', 'ebitda']):
                            metric_name = f"Branch {first_col_val}"
                
                metric_data = {
                    'metric': metric_name,
                    'values': {},
                    'row_index': row_idx
                }
                
                # Add the values
                for time_period, value in zip(time_periods, numeric_values):
                    metric_data['values'][time_period] = value
                
                financial_data[metric_name] = metric_data
        
        return financial_data
    
    def _create_embeddings(self, data):
        """Create embeddings for semantic search with enhanced text representation."""
        if not data:
            return []
        
        # Convert data to text representation
        text_chunks = []
        
        # Add headers
        if data and len(data) > 0:
            headers = [str(cell) for cell in data[0] if cell]
            if headers:
                text_chunks.append("Headers: " + " | ".join(headers))
        
        # Add first column labels (financial metrics)
        for row_idx, row in enumerate(data[1:11]):  # First 10 data rows
            if row and len(row) > 0 and row[0]:
                label = str(row[0]).strip()
                if label and label.lower() not in ['', 'nan', 'none']:
                    text_chunks.append(f"Row {row_idx + 1}: {label}")
        
        # Add sample values for context
        if data and len(data) > 1 and len(data[0]) > 1:
            for col_idx in range(1, min(6, len(data[0]))):  # First 5 columns
                col_values = []
                for row_idx in range(1, min(6, len(data))):  # First 5 rows
                    if row_idx < len(data) and col_idx < len(data[row_idx]):
                        cell_value = data[row_idx][col_idx]
                        if cell_value is not None and str(cell_value).strip():
                            col_values.append(str(cell_value))
                
                if col_values:
                    text_chunks.append(f"Column {col_idx}: " + " | ".join(col_values[:3]))
        
        # Create embeddings
        if text_chunks:
            embeddings = self.model.encode(text_chunks)
            return embeddings.tolist()
        
        return []
